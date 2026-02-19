import asyncio
import os
from datetime import datetime
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import (
    Message, ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery
)
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from openpyxl import Workbook, load_workbook

TOKEN = "8257974390:AAFmP_B9DnyM1sEns62RnmoaxOkXb02qolw"

bot = Bot(token=TOKEN)
dp = Dispatcher()

# ===================== ИНИЦИАЛИЗАЦИЯ =====================

def init_files():
    if not os.path.exists("users.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(["telegram_id", "full_name", "role", "district"])
        wb.save("users.xlsx")

    if not os.path.exists("requests.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "id", "date", "chief_id", "district", "text",
            "deadline", "supply_id", "status",
            "status_date", "reminder_sent", "document_file"
        ])
        wb.save("requests.xlsx")

    if not os.path.exists("documents"):
        os.makedirs("documents")

# ===================== FSM =====================

class Registration(StatesGroup):
    choose_role = State()
    choose_district = State()

class NewRequest(StatesGroup):
    waiting_text = State()
    waiting_deadline = State()

class CloseRequest(StatesGroup):
    waiting_document = State()

class ReportState(StatesGroup):
    waiting_period = State()

# ===================== КЛАВИАТУРЫ =====================

role_keyboard = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="Снабженец")],
              [KeyboardButton(text="Представитель участка")]],
    resize_keyboard=True
)

district_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Кадников")],
        [KeyboardButton(text="Тотьма")],
        [KeyboardButton(text="Нюксеница")],
        [KeyboardButton(text="Шарья")],
        [KeyboardButton(text="Сямжа")],
        [KeyboardButton(text="Верховажье")]
    ],
    resize_keyboard=True
)

chief_menu = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📝 Новая заявка")],
        [KeyboardButton(text="📦 Мои заявки")]
    ],
    resize_keyboard=True
)

supply_menu = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📦 Мои заявки")],
        [KeyboardButton(text="📊 Выгрузить отчёт")]
    ],
    resize_keyboard=True
)

# ===================== РАБОТА С ПОЛЬЗОВАТЕЛЯМИ =====================

def save_user(user_id, full_name, role, district=None):
    wb = load_workbook("users.xlsx")
    ws = wb.active
    ws.append([user_id, full_name, role, district])
    wb.save("users.xlsx")

def get_user(user_id):
    wb = load_workbook("users.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            return {"full_name": row[1], "role": row[2], "district": row[3]}
    return None

def get_supplies():
    wb = load_workbook("users.xlsx")
    ws = wb.active
    return [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[2] == "supply"]

# ===================== START =====================

@dp.message(Command("start"))
async def start_handler(message: Message, state: FSMContext):
    init_files()
    user = get_user(message.from_user.id)

    if user:
        if user["role"] == "chief":
            await message.answer("Вы уже зарегистрированы.", reply_markup=chief_menu)
        else:
            await message.answer("Вы уже зарегистрированы.", reply_markup=supply_menu)
        return

    await state.set_state(Registration.choose_role)
    await message.answer("Выберите роль:", reply_markup=role_keyboard)

# ===================== РЕГИСТРАЦИЯ =====================

@dp.message(Registration.choose_role)
async def process_role(message: Message, state: FSMContext):
    if message.text == "Снабженец":
        save_user(message.from_user.id, message.from_user.full_name, "supply")
        await message.answer("Вы зарегистрированы как снабженец.", reply_markup=supply_menu)
        await state.clear()

    elif message.text == "Представитель участка":
        await state.set_state(Registration.choose_district)
        await message.answer("Выберите участок:", reply_markup=district_keyboard)

@dp.message(Registration.choose_district)
async def process_district(message: Message, state: FSMContext):
    save_user(message.from_user.id, message.from_user.full_name, "chief", message.text)
    await message.answer("Регистрация завершена.", reply_markup=chief_menu)
    await state.clear()

# ===================== СОЗДАНИЕ ЗАЯВКИ =====================

@dp.message(F.text == "📝 Новая заявка")
async def new_request(message: Message, state: FSMContext):
    user = get_user(message.from_user.id)
    if user["role"] != "chief":
        return
    await state.set_state(NewRequest.waiting_text)
    await message.answer("Введите текст заявки:")

@dp.message(NewRequest.waiting_text)
async def get_deadline(message: Message, state: FSMContext):
    await state.update_data(text=message.text)
    await state.set_state(NewRequest.waiting_deadline)
    await message.answer("Введите дедлайн в формате ДД.ММ.ГГГГ")

@dp.message(NewRequest.waiting_deadline)
async def save_request(message: Message, state: FSMContext):
    try:
        datetime.strptime(message.text, "%d.%m.%Y")
    except:
        await message.answer("Неверный формат даты.")
        return

    data = await state.get_data()
    user = get_user(message.from_user.id)

    wb = load_workbook("requests.xlsx")
    ws = wb.active
    new_id = ws.max_row

    ws.append([
        new_id,
        datetime.now().strftime("%d.%m.%Y %H:%M"),
        message.from_user.id,
        user["district"],
        data["text"],
        message.text,
        None,
        "Новая",
        None,
        None,
        None
    ])

    wb.save("requests.xlsx")

    # уведомление снабженцев
    for supply_id in get_supplies():
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="В работу", callback_data=f"work_{new_id}")]
        ])
        await bot.send_message(
            supply_id,
            f"Новая заявка №{new_id}\n{data['text']}\nСрок: {message.text}",
            reply_markup=kb
        )

    await message.answer("Заявка создана.", reply_markup=chief_menu)
    await state.clear()

# ===================== В РАБОТУ =====================

@dp.callback_query(F.data.startswith("work_"))
async def take_in_work(callback: CallbackQuery):
    request_id = int(callback.data.split("_")[1])
    wb = load_workbook("requests.xlsx")
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == request_id:
            row[6].value = callback.from_user.id
            row[7].value = "В работе"
            row[8].value = datetime.now().strftime("%d.%m.%Y %H:%M")
            break

    wb.save("requests.xlsx")

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Закуплено", callback_data=f"close_{request_id}")]
    ])

    await callback.message.edit_text(
        f"Заявка №{request_id} взята в работу.",
        reply_markup=kb
    )

# ===================== ЗАКРЫТИЕ С ДОКУМЕНТОМ =====================

@dp.callback_query(F.data.startswith("close_"))
async def start_close(callback: CallbackQuery, state: FSMContext):
    request_id = int(callback.data.split("_")[1])
    await state.update_data(request_id=request_id)
    await state.set_state(CloseRequest.waiting_document)
    await callback.message.answer("Пришлите УПД или счёт файлом.")

@dp.message(CloseRequest.waiting_document)
async def process_document(message: Message, state: FSMContext):
    if not message.document:
        await message.answer("Нужно отправить документ.")
        return

    data = await state.get_data()
    request_id = data["request_id"]
    file = await bot.get_file(message.document.file_id)
    file_name = f"{request_id}_{message.document.file_name}"
    save_path = os.path.join("documents", file_name)
    await bot.download_file(file.file_path, save_path)

    wb = load_workbook("requests.xlsx")
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[0].value == request_id:
            row[7].value = "Закуплено"
            row[8].value = datetime.now().strftime("%d.%m.%Y %H:%M")
            row[10].value = file_name
            break

    wb.save("requests.xlsx")

    await message.answer("Заявка закрыта.")
    await state.clear()

# ===================== ЗАПУСК =====================

async def main():
    init_files()
    await dp.start_polling(bot)

if name == "main":
    asyncio.run(main())

